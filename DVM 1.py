
import openpyxl
import datetime
from openpyxl import Workbook

#Please make a excel file named Library.xlsx before running the code and mention it's path wherever needed. 
#Also I am extremely sorry to have been unable to stick to the guidelines, due to undue circumstances I had to take leave from campus and haven't had much time to properly explore OOPS and logging and was unable to implement them properly.

excelsheet = openpyxl.load_workbook('Library.xlsx')


ex1=excelsheet['New Title']

ex1['A1']='BOOK NAME'
ex1['B1']='AUTHOR'
ex1['C1']='ISBN'
ex1['D1']='GENRE'
ex1['E1']='Availability'
ex1['F1']='Serial No.'

data=[]
for f in ex1.iter_rows(min_row=2, values_only=True):
    data.append(f)
    



i=0


class book():
    def __init__(self,name,author,ISBN,book_shelf) :
        self.name=name
        self.author=author
        self.ISBN=ISBN
        self.bookshelf=book_shelf
        self.status=True
        global i
        data.append((self.name,self.author,self.ISBN,self.bookshelf))
        i=i+1
        
        ex1.append((self.name,self.author,self.ISBN,self.bookshelf,self.status,i))
        excelsheet.save('Library.xlsx') 

  
        
Genre_list=[]

class shelf():
    def __init__(self):
        cells=ex1['D2':'D{}'.format(count+1)]
        
        global dict
        dict={}
        
        for f in ex1.iter_rows(min_row=2,min_col=4,max_row=count+1,max_col=4,values_only=True):
            
               print(f)    
               for b in f:         
                if b not in Genre_list:
                    Genre_list.append(b)
                else:
                    continue
               
    
                     
        for k in Genre_list:
            list=[]
            j=0 
            for f in ex1.iter_rows(min_row=2,min_col=4,max_row=count+1,max_col=4,values_only=True):
                j=j+1
                for b in f :   
                    if b==k:
                        list.append(j)
               
                dict[k]=list
                

    def show_catalog(self,genre):
     
        print('-------------')
        print(genre,':-')
        print('-----------------')
        j=0
        for k in dict[genre]:

         print(ex1['A{}'.format(dict[genre][j]+1)].value)
         print('----------------')
         j=j+1
        


   
    def get_books_count(self,genre):
        c=0
        for count in dict[genre]:
            c=c+1
        print('The number of Books of ',genre,' are ',c)
        print('--------------------------------')


    def populate_book(self):
        pass



def search_book(search_name):
        book_found=False
        for f in ex1.iter_rows(values_only=True):
           
            name,author,ISBN,genre,availability,srno=f
            if search_name==name:
                print('Book Found \n','--------------------------------')
                

                book_found=True
                return f
            else:
                continue
        if book_found ==False:
            print('BOOK NOT AVAILABLE IN LIBRARY')
            return ('No Book',False,False,False,False,False)
              

def borrow_book(book_name):
        a,b,c,d,availability,e=search_book(book_name)
        
        
        if availability:
            confirmation=str(input('The book is available. Confirm to borrow....Y/N ',))
            if confirmation=='Y':
                availability=False
                ex1['E{}'.format(e+1)]=False
                print('---------------------------')
                print('The Book has been issued at ',datetime.datetime.now())
                print('----------------------------')
            elif confirmation=='N':
                print('----------------------------')
                print('OK')
                print('-----------------------------')
            else:
                print('-----------------------------')
                print('Please enter valid commands')



def return_book(self):
    a,b,c,d,availability,e=search_book(self)
    
    if a!='No Book':
     if not availability:
            confirmation=str(input('The book is not in library.Do you wish to return....Y/N ',))
            if confirmation=='Y':
                availability=True
                ex1['E{}'.format(e+1)]=True
                print('The Book has been returned at ',datetime.datetime.now())
            elif confirmation=='N':
                print('OK')
            else:
                print('Please enter valid commands')
    else: 
        print('Book not available in library register. Ask librarian to add book')


def reserve_book(book_name):
        a,b,c,d,availability,e=search_book(book_name)
       
        if availability:
            confirmation=str(input('The book is available. Confirm to reserve....Y/N ',))
            print('-------------------------------')
            if confirmation=='Y':
                availability=False
                ex1['E{}'.format(e+1)]=False
                print('The Book has been reserved at ',datetime.datetime.now())
                print('------------------------------')
            elif confirmation=='N':
                print('OK')
                print('------------------------------')
            else:
                print('Please enter valid commands')
                print('-------------------------------')

        

A=True
while A:

    print('Enter User type Librarian or Student L/S')
    user_type=str(input())

    if user_type=='L':
        print('----------------------------------------')
        print('Librarian')
        print('----------------------------------------')
        All_access=True
        A=False
    elif user_type=='S':
        print('----------------------------------------')
        print('Student')
        print('----------------------------------------')
        All_access=False
        A=False
    else:
        print('----------------------------------------')
        print('invalid user type, Input L or S')
        print('----------------------------------------')



A=True
while A:
  count=len(data)
  if All_access:  
    print('Choose operation to be performed')
    print('----------------------------------------')
    print('Add book                - 1')
    print('Borrow book             - 2')
    print('Reserve book            - 3')
    print('Return book             - 4')
    print('Show books of a genre   - 5')
    print('Exit                    - 6')
    print('----------------------------------------')
    q=str(input())
    if q=='1':
        name=str(input('Enter Book Name '))
        author=str(input('Enter Author Name '))
        ISBN=str(input('Enter ISBN number '))
        Genre=str(input('Enter Book Genre '))
        book(name,author,ISBN,Genre)
        print('----------------------------------------')
        print('Book successfullly added')
        print('----------------------------------------')
    elif q=='2':
        name=str(input('Enter Book name to borrow '))
        borrow_book(name)

    elif q=='3':
        name=str(input('Enter Book name to reserve book'))
        reserve_book(name)
    elif q=='4':
        name=str(input('Enter Book name to return '))
        return_book(name)

    elif q=='5':
        genre=str(input('Enter Genre name '))
        genre1=shelf()
        genre1.show_catalog(genre)
        genre1.get_books_count(genre)

    elif q=='6':
        A=False
  else:
    print('Choose operation to be performed')
    print('----------------------------------------')
    
    print('Borrow book             - 1')
    print('Reserve book            - 2')
    print('Return book             - 3')
    print('Show books of a genre   - 4')
    print('Exit                    - 5')
    print('----------------------------------------')
    q=str(input())
    if q=='1':
        name=str(input('Enter Book name to borrow '))
        borrow_book(name)

    elif q=='2':
        name=str(input('Enter Book name to reserve book'))
        reserve_book(name)
    elif q=='3':
        name=str(input('Enter Book name to return '))
        return_book(name)

    elif q=='4':
        genre=str(input('Enter Genre name '))
        genre1=shelf()
        genre1.show_catalog(genre)
        genre1.get_books_count(genre)

    elif q=='5':
        A=False

    















excelsheet.save('Library.xlsx')      